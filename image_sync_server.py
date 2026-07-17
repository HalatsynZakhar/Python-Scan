from __future__ import annotations

import argparse
import io
import json
import os
import secrets
import threading
import time
from contextlib import asynccontextmanager
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any

import requests
import uvicorn
from fastapi import FastAPI, HTTPException, Request, Response
from fastapi.responses import HTMLResponse
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from starlette.datastructures import UploadFile as StarletteUploadFile
from watchdog.events import FileSystemEvent, FileSystemEventHandler
from watchdog.observers import Observer

import images_xml
from horoshop_sync import (
    CatalogIndex,
    HoroshopClient,
    HoroshopError,
    HoroshopSettings,
    build_clear_product,
    build_import_product,
    chunked,
    force_append_upload,
    current_epoch,
    import_article_succeeded,
    import_log_by_article,
    load_horoshop_settings,
    load_xml_products,
    normalize_article,
    read_raw_config,
    with_runtime_credentials,
    XmlProduct,
)
from images_xml import (
    DEFAULT_CONFIG_FILE,
    EVENT_TYPE_LABELS,
    WATCH_IGNORED_SUFFIXES,
    build_xml,
    configure_console_encoding,
    configure_logging,
    initialize_reliable_clock,
    load_config,
    log,
    parse_filename,
)


STATE_SCHEMA_VERSION = 2
STATE_HISTORY_LIMIT = 300
JOB_TTL_SECONDS = 60 * 60
CATALOG_CACHE_MAX_ITEMS = 200_000
LOCAL_DATA_DIR = Path(__file__).resolve().parent / "data"

CONFIG_FILE = DEFAULT_CONFIG_FILE
RAW_CONFIG: dict[str, Any] = {}
XML_CONFIG: dict[str, Any] = {}
HOROSHOP_SETTINGS: HoroshopSettings | None = None
SERVER_SETTINGS: dict[str, Any] = {}
STATE: "SyncState | None" = None
STATE_LOCK = threading.Lock()
SYNC_LOCK = threading.Lock()
OBSERVER: Observer | None = None
JOBS: dict[str, dict[str, Any]] = {}
JOBS_LOCK = threading.Lock()


@dataclass(frozen=True)
class XmlRequest:
    article: str
    brand: str = ""


def normalize_brand(value: Any) -> str:
    return str(value or "").strip()


def state_item_key(article: str, brand: str = "") -> str:
    """Keep legacy article-only keys for products without a brand."""
    article = normalize_article(article)
    brand = normalize_brand(brand)
    return article if not brand else f"{article}\x1f{brand}"


def request_key(article: str, brand: str = "") -> tuple[str, str]:
    return normalize_article(article), normalize_brand(brand)


def format_available_brands(products: list[XmlProduct], article: str) -> str:
    brands = []
    seen: set[str] = set()
    for product in products:
        if product.article != article or product.brand in seen:
            continue
        seen.add(product.brand)
        brands.append(product.brand or "(порожній бренд)")
    return ", ".join(brands) or "немає"


def select_xml_product(
    xml_products: list[XmlProduct],
    article: str,
    brand: str = "",
) -> tuple[XmlProduct | None, str]:
    article = normalize_article(article)
    brand = normalize_brand(brand)
    candidates = [product for product in xml_products if product.article == article]
    if not candidates:
        return None, f"У XML не знайдено фото для артикула {article}."
    if not brand:
        return candidates[0], ""

    for product in candidates:
        if product.brand == brand:
            return product, ""
    folded_brand = brand.casefold()
    for product in candidates:
        if product.brand.casefold() == folded_brand:
            return product, ""
    return (
        None,
        f"Не знайдено бренд '{brand}' для артикула {article}. "
        f"Доступні бренди: {format_available_brands(xml_products, article)}.",
    )


def load_server_settings(raw: dict[str, Any], xml_config: dict[str, Any]) -> dict[str, Any]:
    server = raw.get("server") or {}
    if not isinstance(server, dict):
        raise ValueError("Секція server у config.json повинна бути об'єктом.")

    state_file = server.get("state_file")
    if state_file:
        resolved_state_file = Path(str(state_file)).expanduser()
    else:
        resolved_state_file = LOCAL_DATA_DIR / "horoshop_sync_state.json"

    try:
        resolved_state_file.resolve().relative_to(xml_config["output_dir"].resolve())
    except ValueError:
        pass
    else:
        raise ValueError(
            "server.state_file не можна зберігати в output_dir. "
            "У публічній папці мають бути лише images_export.xml та звичайний журнал."
        )

    return {
        "enabled": bool(server.get("enabled", False)),
        "host": str(server.get("host", "0.0.0.0")),
        "port": int(server.get("port", 8092)),
        "state_file": resolved_state_file,
    }


def timestamp() -> str:
    return images_xml.now_kyiv().isoformat(timespec="seconds")


class SyncState:
    def __init__(self, state_file: Path):
        self.state_file = state_file
        self.dirty: dict[str, dict[str, Any]] = {}
        self.history: list[dict[str, Any]] = []
        self.archive: list[dict[str, Any]] = []
        self.catalog_products: list[dict[str, Any]] = []
        self.catalog_updated_at = ""
        self.load()

    def load(self) -> None:
        if not self.state_file.exists():
            return
        with self.state_file.open("r", encoding="utf-8") as file:
            data = json.load(file)
        if not isinstance(data, dict):
            return
        dirty = data.get("dirty", {})
        history = data.get("history", [])
        archive = data.get("archive", [])
        catalog = data.get("catalog", {})
        if isinstance(dirty, dict):
            self.dirty = {}
            for article, item in dirty.items():
                if not isinstance(item, dict):
                    continue
                normalized_article = normalize_article(item.get("article") or article)
                brand = normalize_brand(item.get("brand"))
                if not normalized_article:
                    continue
                normalized_item = dict(item)
                normalized_item["article"] = normalized_article
                normalized_item["brand"] = brand
                self.dirty[state_item_key(normalized_article, brand)] = normalized_item
        if isinstance(history, list):
            self.history = [item for item in history if isinstance(item, dict)][
                -STATE_HISTORY_LIMIT:
            ]
        if isinstance(archive, list):
            self.archive = [item for item in archive if isinstance(item, dict)][
                -STATE_HISTORY_LIMIT:
            ]
        if isinstance(catalog, dict):
            products = catalog.get("products", [])
            if isinstance(products, list):
                self.catalog_products = [
                    item for item in products if isinstance(item, dict)
                ][:CATALOG_CACHE_MAX_ITEMS]
            self.catalog_updated_at = normalize_article(catalog.get("updated_at"))

    def save(self) -> None:
        self.state_file.parent.mkdir(parents=True, exist_ok=True)
        payload = {
            "schema_version": STATE_SCHEMA_VERSION,
            "updated_at": timestamp(),
            "dirty": self.dirty,
            "history": self.history[-STATE_HISTORY_LIMIT:],
            "archive": self.archive[-STATE_HISTORY_LIMIT:],
            "catalog": {
                "updated_at": self.catalog_updated_at,
                "products_count": len(self.catalog_products),
                "products": self.catalog_products,
            },
        }
        temp_file = self.state_file.with_suffix(f"{self.state_file.suffix}.tmp")
        with temp_file.open("w", encoding="utf-8") as file:
            json.dump(payload, file, ensure_ascii=False, indent=2)
        os.replace(temp_file, self.state_file)

    def mark_dirty(
        self,
        article: str,
        event_type: str,
        image_count: int | None,
        message: str | None = None,
        brand: str = "",
    ) -> None:
        article = normalize_article(article)
        brand = normalize_brand(brand)
        if not article:
            return
        now = timestamp()
        key = state_item_key(article, brand)
        existing = self.dirty.get(key, {})
        self.dirty[key] = {
            "article": article,
            "brand": brand,
            "first_seen_at": existing.get("first_seen_at") or now,
            "updated_at": now,
            "event_type": event_type,
            "events": int(existing.get("events", 0)) + 1,
            "status": "dirty",
            "message": message or "Очікує оновлення в Хорошопі.",
            "image_count": image_count,
        }

    def mark_result(
        self,
        *,
        article: str,
        status: str,
        message: str,
        remote_article: str = "",
        image_count: int | None = None,
        clear_dirty: bool = False,
        brand: str = "",
    ) -> None:
        article = normalize_article(article)
        brand = normalize_brand(brand)
        if not article:
            return
        key = state_item_key(article, brand)
        item = self.dirty.get(key, {"article": article, "brand": brand})
        item.update(
            {
                "brand": brand,
                "updated_at": timestamp(),
                "status": status,
                "message": message,
                "remote_article": remote_article,
                "image_count": image_count,
            }
        )
        self.history.append(dict(item))
        self.history = self.history[-STATE_HISTORY_LIMIT:]
        if clear_dirty:
            self.dirty.pop(key, None)
        else:
            self.dirty[key] = item

    def clear_dirty(self) -> None:
        self.dirty.clear()

    def skip_all_dirty(self) -> int:
        if not self.dirty:
            return 0
        moved = 0
        for article in sorted(list(self.dirty), key=str.casefold):
            item = dict(self.dirty.pop(article))
            item.update(
                {
                    "updated_at": timestamp(),
                    "status": "skipped",
                    "message": "Пропущено разом з усією чергою.",
                }
            )
            self.history.append(item)
            moved += 1
        self.history = self.history[-STATE_HISTORY_LIMIT:]
        return moved

    def skip_dirty(self, article: str, brand: str = "") -> bool:
        article = normalize_article(article)
        key = state_item_key(article, brand)
        if not article or key not in self.dirty:
            return False
        item = dict(self.dirty.pop(key))
        item.update(
            {
                "updated_at": timestamp(),
                "status": "skipped",
                "message": "Пропущено вручну.",
            }
        )
        self.history.append(item)
        self.history = self.history[-STATE_HISTORY_LIMIT:]
        return True

    def set_catalog_products(self, products: list[dict[str, Any]]) -> None:
        self.catalog_products = [
            item for item in products if isinstance(item, dict)
        ][:CATALOG_CACHE_MAX_ITEMS]
        self.catalog_updated_at = timestamp()

    def catalog_meta(self) -> dict[str, Any]:
        return {
            "updated_at": self.catalog_updated_at,
            "products_count": len(self.catalog_products),
            "has_cache": bool(self.catalog_products),
            "local_path": str(self.state_file),
        }

    def archive_history(self) -> int:
        if not self.history:
            return 0
        moved = len(self.history)
        archived_at = timestamp()
        self.archive.extend(
            {
                **item,
                "archived_at": archived_at,
            }
            for item in self.history
        )
        self.archive = self.archive[-STATE_HISTORY_LIMIT:]
        self.history.clear()
        return moved

    def archive_history_item(
        self,
        article: str,
        updated_at: str = "",
        brand: str = "",
    ) -> bool:
        article = normalize_article(article)
        brand = normalize_brand(brand)
        if not article:
            return False
        for index, item in enumerate(self.history):
            if normalize_article(item.get("article")) != article:
                continue
            if normalize_brand(item.get("brand")) != brand:
                continue
            item_updated_at = str(item.get("updated_at") or item.get("first_seen_at") or "")
            if updated_at and item_updated_at != updated_at:
                continue
            archived = dict(self.history.pop(index))
            archived["archived_at"] = timestamp()
            self.archive.append(archived)
            self.archive = self.archive[-STATE_HISTORY_LIMIT:]
            return True
        return False


def get_state() -> SyncState:
    if STATE is None:
        raise RuntimeError("Стан синхронізації ще не ініціалізовано.")
    return STATE


def cleanup_jobs() -> None:
    cutoff = current_epoch() - JOB_TTL_SECONDS
    with JOBS_LOCK:
        expired = [
            job_id
            for job_id, item in JOBS.items()
            if float(item.get("updated_epoch", 0)) < cutoff
        ]
        for job_id in expired:
            JOBS.pop(job_id, None)


def set_job(job_id: str, **updates: Any) -> None:
    cleanup_jobs()
    with JOBS_LOCK:
        item = JOBS.get(job_id, {})
        item.update(updates)
        item["updated_epoch"] = current_epoch()
        JOBS[job_id] = item


def get_job(job_id: str) -> dict[str, Any]:
    cleanup_jobs()
    with JOBS_LOCK:
        item = dict(JOBS.get(job_id, {}))
    if not item:
        return {"status": "unknown", "percent": 0, "message": "Завдання не знайдено."}
    return item


def protected_json(_: Request) -> None:
    return


def image_articles_from_event(
    event: FileSystemEvent,
    allowed_extensions: set[str],
    images_dir: Path,
) -> set[XmlRequest]:
    paths = [event.src_path]
    destination = getattr(event, "dest_path", "")
    if destination:
        paths.append(destination)

    articles: set[XmlRequest] = set()
    for path in paths:
        file_path = Path(path)
        suffixes = {suffix.lower() for suffix in file_path.suffixes}
        if suffixes & WATCH_IGNORED_SUFFIXES:
            continue
        if file_path.suffix.lower() not in allowed_extensions:
            continue
        article, _ = parse_filename(file_path)
        if article:
            try:
                brand = file_path.relative_to(images_dir).parent.as_posix()
            except ValueError:
                brand = ""
            articles.add(XmlRequest(article, "" if brand == "." else brand))
    return articles


def image_counts(xml_products: list[XmlProduct]) -> dict[tuple[str, str], int]:
    return {
        request_key(product.article, product.brand): len(product.image_urls)
        for product in xml_products
    }


def add_manual_dirty_article(article: str, brand: str = "") -> dict[str, Any]:
    normalized = normalize_article(article)
    requested_brand = normalize_brand(brand)
    if not normalized:
        raise ValueError("Введіть артикул для перевірки.")

    build_xml(XML_CONFIG)
    try:
        xml_products = load_xml_products(XML_CONFIG["output_xml"])
    except OSError as error:
        raise ValueError(f"Не вдалося прочитати XML: {error}") from error

    product, reason = select_xml_product(xml_products, normalized, requested_brand)
    if product is None or not product.image_urls:
        raise ValueError(reason or f"У XML не знайдено фото для артикула {normalized}.")

    with STATE_LOCK:
        state = get_state()
        state.mark_dirty(
            product.article,
            "manual",
            len(product.image_urls),
            brand=product.brand,
        )
        state.save()

    return {
        "ok": True,
        "article": product.article,
        "brand": product.brand,
        "image_count": len(product.image_urls),
        "sample_images": list(product.image_urls[:5]),
    }


def queue_preview_articles(
    *,
    plan: dict[str, Any],
    validation_failed: list[dict[str, str]],
    counts: dict[tuple[str, str], int],
    event_type: str,
) -> dict[str, Any]:
    failed_messages = {
        request_key(item.get("article"), item.get("brand")): str(item.get("message", ""))
        for item in validation_failed
    }
    queued = 0
    with STATE_LOCK:
        state = get_state()
        for item in plan.get("preview", []):
            article = normalize_article(item.get("article"))
            brand = normalize_brand(item.get("brand"))
            if not article:
                continue
            key = request_key(article, brand)
            image_count = counts.get(key, int(item.get("image_count") or 0))
            if key in failed_messages:
                message = f"Перевірено через Excel. Помилка URL: {failed_messages[key]}"
            elif item.get("status") == "ready":
                message = "Перевірено через Excel. Можна оновити точково або разом з чергою."
            else:
                message = f"Перевірено через Excel. {item.get('message') or 'Потрібна увага.'}"
            state.mark_dirty(article, event_type, image_count, message=message, brand=brand)
            queued += 1
        state.save()
    return {"queued": queued}


def credentials_from_form(form: Any) -> dict[str, str]:
    return {
        "login": str(form.get("login", "")).strip(),
        "password": str(form.get("password", "")),
        "token": str(form.get("token", "")).strip(),
    }


def fresh_catalog_from_form(form: Any) -> bool:
    return str(form.get("fresh_catalog", "")).strip().lower() in {
        "1",
        "true",
        "yes",
        "on",
    }


def unique_xml_requests(values: list[XmlRequest]) -> list[XmlRequest]:
    result: list[XmlRequest] = []
    seen: set[tuple[str, str]] = set()
    for value in values:
        article, brand = request_key(value.article, value.brand)
        if not article:
            continue
        key = (article.casefold(), brand.casefold())
        if key in seen:
            continue
        seen.add(key)
        result.append(XmlRequest(article, brand))
    return result


def parse_excel_articles(data: bytes) -> list[XmlRequest]:
    workbook = load_workbook(
        io.BytesIO(data),
        read_only=True,
        data_only=True,
    )
    try:
        worksheet = workbook.worksheets[0]
        values: list[XmlRequest] = []
        for row in worksheet.iter_rows(values_only=True):
            if not row:
                continue
            value = row[0]
            if value is None:
                continue
            text = normalize_article(value)
            if not text:
                continue
            if text.casefold() in {
                "article",
                "артикул",
                "артикул для відображення",
                "артикул для отображения",
            }:
                continue
            brand = normalize_brand(row[1]) if len(row) > 1 else ""
            values.append(XmlRequest(text, brand))
        return unique_xml_requests(values)
    finally:
        workbook.close()


def build_excel_template() -> bytes:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Імпорт"
    worksheet.append(["Артикул", "Бренд"])
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = "A1:B1"
    worksheet.column_dimensions["A"].width = 24
    worksheet.column_dimensions["B"].width = 32

    header_fill = PatternFill("solid", fgColor="166534")
    for cell in worksheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for row in range(2, 102):
        worksheet.cell(row=row, column=1)
        worksheet.cell(row=row, column=2)

    output = io.BytesIO()
    workbook.save(output)
    workbook.close()
    return output.getvalue()


def load_or_refresh_catalog(
    *,
    client: HoroshopClient,
    force_refresh: bool,
    progress: Any | None = None,
) -> tuple[CatalogIndex, dict[str, Any]]:
    with STATE_LOCK:
        state = get_state()
        cached_products = list(state.catalog_products)
        cached_meta = state.catalog_meta()

    if cached_products and not force_refresh:
        return CatalogIndex.from_raw(cached_products), {
            **cached_meta,
            "source": "cache",
        }

    if progress:
        progress(0, "Завантаження свіжого каталогу Хорошопа...")
    products = client.export_catalog(progress=progress)
    with STATE_LOCK:
        state = get_state()
        state.set_catalog_products(products)
        state.save()
        meta = state.catalog_meta()
    return CatalogIndex.from_raw(products), {**meta, "source": "fresh"}


def catalog_age_seconds(updated_at: str) -> int | None:
    if not updated_at:
        return None
    try:
        parsed = datetime.fromisoformat(updated_at)
    except ValueError:
        return None
    return max(0, int((images_xml.now_kyiv() - parsed).total_seconds()))


def resolve_target_articles(
    *,
    mode: str,
    xml_products: list[XmlProduct],
    requested_articles: list[XmlRequest] | None,
) -> list[XmlRequest]:
    with STATE_LOCK:
        dirty_articles = [
            XmlRequest(
                normalize_article(item.get("article")),
                normalize_brand(item.get("brand")),
            )
            for item in get_state().dirty.values()
        ]

    if mode == "dirty":
        target_articles = requested_articles or dirty_articles
    elif mode in {"article", "excel"}:
        target_articles = requested_articles or []
    elif mode == "all":
        target_articles = [
            XmlRequest(product.article, product.brand) for product in xml_products
        ]
    else:
        raise ValueError("Невідомий режим синхронізації.")
    return unique_xml_requests(target_articles)


def prepare_import_plan(
    *,
    settings: HoroshopSettings,
    catalog: CatalogIndex,
    xml_products: list[XmlProduct],
    target_articles: list[XmlRequest],
) -> dict[str, Any]:
    prepared: list[dict[str, Any]] = []
    skipped: list[dict[str, str]] = []
    preview: list[dict[str, Any]] = []

    for request in target_articles:
        product, selection_error = select_xml_product(
            xml_products,
            request.article,
            request.brand,
        )
        local_article = request.article
        brand = request.brand
        image_urls = list(product.image_urls) if product else []
        if product is not None:
            local_article = product.article
            brand = product.brand
        match = catalog.match(local_article)
        item, reason = build_import_product(
            match=match,
            image_urls=image_urls,
            settings=settings,
        )
        if product is None:
            item = None
            reason = selection_error
        remote_article = item["article"] if item is not None else ""
        preview_item = {
            "article": local_article,
            "brand": brand,
            "remote_article": remote_article,
            "status": "ready" if item is not None else "skipped",
            "message": reason,
            "image_count": len(image_urls),
            "sample_images": image_urls[:5],
        }
        preview.append(preview_item)
        if item is None:
            skipped.append({"article": local_article, "brand": brand, "message": reason})
            continue
        prepared.append(
            {
                "local_article": local_article,
                "brand": brand,
                "remote_article": remote_article,
                "item": item,
                "has_images": bool(item.get(settings.image_field, {}).get("links")),
                "image_urls": image_urls,
            }
        )

    return {
        "prepared": prepared,
        "skipped": skipped,
        "preview": preview,
    }


def validate_image_url(
    url: str,
    session: requests.Session,
    *,
    timeout: int = 10,
    max_bytes: int = 5 * 1024 * 1024,
) -> tuple[bool, str]:
    try:
        response = session.head(url, allow_redirects=True, timeout=timeout)
        if response.status_code in {405, 403}:
            response = session.get(url, stream=True, timeout=timeout)
        response.raise_for_status()
    except requests.RequestException as error:
        return False, f"URL недоступний: {error}"

    content_type = response.headers.get("content-type", "").split(";", 1)[0].strip().lower()
    if content_type and content_type not in {"image/jpeg", "image/png", "image/gif"}:
        return False, f"Некоректний MIME: {content_type}"

    content_length = response.headers.get("content-length")
    if content_length:
        try:
            if int(content_length) > max_bytes:
                return False, "Файл більший за 5 МБ"
        except ValueError:
            pass
    return True, "OK"


def validate_prepared_images(
    prepared: list[dict[str, Any]],
    *,
    progress: Any | None = None,
) -> tuple[list[dict[str, Any]], list[dict[str, str]], dict[str, Any]]:
    valid: list[dict[str, Any]] = []
    failed: list[dict[str, str]] = []
    checked = 0
    session = requests.Session()

    for item in prepared:
        article = item["local_article"]
        errors: list[str] = []
        for url in item.get("image_urls", []):
            checked += 1
            if progress and checked % 10 == 0:
                progress(checked, f"Перевірено URL фото: {checked}")
            ok, message = validate_image_url(url, session)
            if not ok:
                errors.append(f"{url}: {message}")
                if len(errors) >= 3:
                    break
        if errors:
            failed.append(
                {
                    "article": article,
                    "brand": str(item.get("brand", "")),
                    "message": " | ".join(errors),
                }
            )
        else:
            valid.append(item)

    return valid, failed, {"checked_urls": checked}


class SyncChangeHandler(FileSystemEventHandler):
    def __init__(self, config: dict[str, Any]):
        super().__init__()
        self.config = config
        self.allowed_extensions: set[str] = config["allowed_extensions"]
        self._build_lock = threading.Lock()

    def _handle(self, event: FileSystemEvent) -> None:
        articles = image_articles_from_event(
            event,
            self.allowed_extensions,
            self.config["images_dir"],
        )
        if not articles and not event.is_directory:
            return
        if not self._build_lock.acquire(blocking=False):
            return

        try:
            event_type = EVENT_TYPE_LABELS.get(event.event_type, event.event_type)
            build_xml(self.config)
            xml_products = load_xml_products(self.config["output_xml"])
            counts = image_counts(xml_products)

            valid_articles = {
                request
                for request in articles
                if request_key(request.article, request.brand) in counts
            }
            if valid_articles:
                with STATE_LOCK:
                    state = get_state()
                    for request in valid_articles:
                        state.mark_dirty(
                            request.article,
                            event_type,
                            counts.get(request_key(request.article, request.brand), 0),
                            brand=request.brand,
                        )
                    state.save()
                log(
                    "Зафіксовано зміни для артикулів: "
                    + ", ".join(
                        f"{request.article} [{request.brand or '-'}]"
                        for request in sorted(
                            valid_articles,
                            key=lambda item: (item.article.casefold(), item.brand.casefold()),
                        )
                    )
                )
        except Exception as error:
            log(f"Помилка оновлення XML/черги: {error}")
        finally:
            self._build_lock.release()

    def on_created(self, event: FileSystemEvent) -> None:
        self._handle(event)

    def on_modified(self, event: FileSystemEvent) -> None:
        if not event.is_directory:
            self._handle(event)

    def on_deleted(self, event: FileSystemEvent) -> None:
        self._handle(event)

    def on_moved(self, event: FileSystemEvent) -> None:
        self._handle(event)


def start_observer() -> Observer:
    build_xml(XML_CONFIG)
    observer = Observer()
    observer.schedule(SyncChangeHandler(XML_CONFIG), str(XML_CONFIG["images_dir"]), recursive=True)
    observer.start()
    log(f"Веб-сервер і відстеження змін запущено: {XML_CONFIG['images_dir']}")
    return observer


def sync_job(
    job_id: str,
    mode: str,
    credentials: dict[str, Any],
    requested_articles: list[XmlRequest] | None = None,
    fresh_catalog: bool = False,
) -> None:
    base_settings = HOROSHOP_SETTINGS
    if base_settings is None:
        set_job(
            job_id,
            status="error",
            percent=100,
            message="Налаштування Хорошопа не ініціалізовано.",
        )
        return

    if not SYNC_LOCK.acquire(blocking=False):
        set_job(
            job_id,
            status="error",
            percent=100,
            message="Інша синхронізація вже виконується.",
        )
        return

    try:
        settings = with_runtime_credentials(base_settings, credentials)
        set_job(job_id, status="running", percent=2, message="Оновлення XML...")
        build_xml(XML_CONFIG)
        xml_products = load_xml_products(XML_CONFIG["output_xml"])
        counts = image_counts(xml_products)

        target_articles = resolve_target_articles(
            mode=mode,
            xml_products=xml_products,
            requested_articles=requested_articles,
        )

        if not target_articles:
            set_job(job_id, status="done", percent=100, message="Немає артикулів для оновлення.")
            return

        client = HoroshopClient(settings)

        def export_progress(count: int, message: str) -> None:
            set_job(job_id, status="running", percent=10, message=message, exported=count)

        set_job(
            job_id,
            status="running",
            percent=8,
            message=(
                "Завантаження каталогу Хорошопа..."
                if fresh_catalog
                else "Підготовка каталогу Хорошопа..."
            ),
        )
        catalog, catalog_meta = load_or_refresh_catalog(
            client=client,
            force_refresh=fresh_catalog,
            progress=export_progress,
        )
        set_job(
            job_id,
            status="running",
            percent=25,
            message="Зіставлення артикулів...",
            catalog_products=len(catalog.products),
            catalog_source=catalog_meta.get("source"),
            catalog_updated_at=catalog_meta.get("updated_at"),
        )

        should_update_state = mode in {"dirty", "article", "excel"}
        plan = prepare_import_plan(
            settings=settings,
            catalog=catalog,
            xml_products=xml_products,
            target_articles=target_articles,
        )
        skipped: list[dict[str, str]] = plan["skipped"]
        prepared_items: list[dict[str, Any]] = plan["prepared"]
        for skipped_item in skipped:
            if should_update_state:
                with STATE_LOCK:
                    state = get_state()
                    state.mark_result(
                        article=skipped_item["article"],
                        brand=str(skipped_item.get("brand", "")),
                        status="skipped",
                        message=skipped_item["message"],
                        image_count=counts.get(
                            request_key(
                                skipped_item["article"],
                                str(skipped_item.get("brand", "")),
                            ),
                            0,
                        ),
                    )
                    state.save()

        def validation_progress(count: int, message: str) -> None:
            set_job(job_id, status="running", percent=30, message=message, checked_urls=count)

        set_job(job_id, status="running", percent=28, message="Перевірка посилань на фото...")
        prepared_items, validation_failed, validation_meta = validate_prepared_images(
            prepared_items,
            progress=validation_progress,
        )
        if validation_failed:
            for failed_item in validation_failed:
                if should_update_state:
                    with STATE_LOCK:
                        state = get_state()
                        state.mark_result(
                            article=failed_item["article"],
                            brand=str(failed_item.get("brand", "")),
                            status="error",
                            message=failed_item["message"],
                            image_count=counts.get(
                                request_key(
                                    failed_item["article"],
                                    str(failed_item.get("brand", "")),
                                ),
                                0,
                            ),
                        )
                        state.save()

        if not prepared_items:
            set_job(
                job_id,
                status="done",
                percent=100,
                message="Немає підготовлених товарів для імпорту.",
                skipped=skipped,
                failed=validation_failed,
                imported=0,
                report={
                    "prepared": 0,
                    "skipped": skipped,
                    "failed": validation_failed,
                    "preview": plan["preview"],
                    **validation_meta,
                },
            )
            return

        imported = 0
        failed: list[dict[str, str]] = []
        batches = chunked(prepared_items, settings.batch_size)
        for index, batch in enumerate(batches, start=1):
            percent = 25 + ((index - 1) / len(batches)) * 70
            set_job(
                job_id,
                status="running",
                percent=round(percent, 1),
                message=f"Імпорт у Хорошоп: пакет {index} із {len(batches)}...",
                imported=imported,
                total=len(prepared_items),
            )
            if settings.override and settings.two_phase_replace:
                clear_items = [
                    build_clear_product(prepared["remote_article"], settings)
                    for prepared in batch
                ]
                clear_response = client.import_products(clear_items)
                clear_status = str(clear_response.get("status", "OK")).upper()
                clear_logs = import_log_by_article(clear_response)
                upload_items = [
                    force_append_upload(prepared["item"], settings)
                    for prepared in batch
                    if prepared["has_images"]
                ]
                response = (
                    client.import_products(upload_items)
                    if upload_items
                    else clear_response
                )
                response_status = str(response.get("status", "OK")).upper()
                logs = import_log_by_article(response)
            else:
                raw_batch = [prepared["item"] for prepared in batch]
                clear_status = ""
                clear_logs = {}
                response = client.import_products(raw_batch)
                response_status = str(response.get("status", "OK")).upper()
                logs = import_log_by_article(response)

            for prepared in batch:
                remote_article = prepared["remote_article"]
                local_article = prepared["local_article"]
                brand = prepared["brand"]
                article_log = logs.get(remote_article, [])
                clear_success = (
                    True
                    if not clear_status
                    else import_article_succeeded(
                        clear_status,
                        clear_logs.get(remote_article, []),
                    )
                )
                if settings.override and settings.two_phase_replace and not prepared["has_images"]:
                    success = clear_success
                    article_log = clear_logs.get(remote_article, [])
                else:
                    success = clear_success and import_article_succeeded(
                        response_status,
                        article_log,
                    )
                if success:
                    imported += 1
                    if should_update_state:
                        with STATE_LOCK:
                            state = get_state()
                            state.mark_result(
                                article=local_article,
                                brand=brand,
                                status="synced",
                                message="Зображення оновлено в Хорошопі.",
                                remote_article=remote_article,
                                image_count=counts.get(
                                    request_key(local_article, brand), 0
                                ),
                                clear_dirty=True,
                            )
                            state.save()
                else:
                    message = "; ".join(
                        str(entry.get("message", entry)) for entry in article_log
                    ) or "; ".join(
                        str(entry.get("message", entry))
                        for entry in clear_logs.get(remote_article, [])
                    ) or f"Статус імпорту: {response_status}"
                    failed.append(
                        {"article": local_article, "brand": brand, "message": message}
                    )
                    if should_update_state:
                        with STATE_LOCK:
                            state = get_state()
                            state.mark_result(
                                article=local_article,
                                brand=brand,
                                status="error",
                                message=message,
                                remote_article=remote_article,
                                image_count=counts.get(
                                    request_key(local_article, brand), 0
                                ),
                            )
                            state.save()

        set_job(
            job_id,
            status="done" if not failed else "warning",
            percent=100,
            message=(
                f"Готово. Оновлено: {imported}; пропущено: {len(skipped)}; "
                f"помилок: {len(failed)}."
            ),
            imported=imported,
            skipped=skipped,
            failed=failed,
            total=len(prepared_items),
            catalog_source=catalog_meta.get("source"),
            catalog_updated_at=catalog_meta.get("updated_at"),
            report={
                "prepared": len(prepared_items),
                "imported": imported,
                "skipped": skipped,
                "failed": [*validation_failed, *failed],
                "preview": plan["preview"],
                "checked_urls": validation_meta.get("checked_urls", 0),
                "catalog_source": catalog_meta.get("source"),
                "catalog_updated_at": catalog_meta.get("updated_at"),
            },
        )
        log(
            f"Синхронізація Хорошоп завершена: mode={mode}, "
            f"updated={imported}, skipped={len(skipped)}, failed={len(failed)}"
        )
    except (HoroshopError, OSError, ValueError) as error:
        log(f"Помилка синхронізації Хорошоп: {error}")
        set_job(job_id, status="error", percent=100, message=str(error))
    finally:
        SYNC_LOCK.release()


def start_sync(
    mode: str,
    credentials: dict[str, Any],
    articles: list[XmlRequest] | None = None,
    fresh_catalog: bool = False,
) -> dict[str, str]:
    job_id = secrets.token_hex(12)
    set_job(job_id, status="queued", percent=0, message="Очікування запуску...")
    thread = threading.Thread(
        target=sync_job,
        args=(job_id, mode, credentials, articles, fresh_catalog),
        daemon=True,
    )
    thread.start()
    return {"job_id": job_id}


def refresh_catalog_job(job_id: str, credentials: dict[str, Any]) -> None:
    base_settings = HOROSHOP_SETTINGS
    if base_settings is None:
        set_job(
            job_id,
            status="error",
            percent=100,
            message="Налаштування Хорошопа не ініціалізовано.",
        )
        return

    if not SYNC_LOCK.acquire(blocking=False):
        set_job(
            job_id,
            status="error",
            percent=100,
            message="Інша операція вже виконується.",
        )
        return

    try:
        settings = with_runtime_credentials(base_settings, credentials)
        client = HoroshopClient(settings)

        def export_progress(count: int, message: str) -> None:
            set_job(job_id, status="running", percent=50, message=message, exported=count)

        set_job(job_id, status="running", percent=5, message="Старт експорту каталогу Хорошопа...")
        catalog, meta = load_or_refresh_catalog(
            client=client,
            force_refresh=True,
            progress=export_progress,
        )
        set_job(
            job_id,
            status="done",
            percent=100,
            message=f"Каталог оновлено: товарів {len(catalog.products)}.",
            catalog_products=len(catalog.products),
            catalog_updated_at=meta.get("updated_at"),
        )
        log(f"Кеш каталогу Хорошопа оновлено: товарів {len(catalog.products)}")
    except (HoroshopError, OSError, ValueError) as error:
        log(f"Помилка оновлення кешу каталогу Хорошопа: {error}")
        set_job(job_id, status="error", percent=100, message=str(error))
    finally:
        SYNC_LOCK.release()


def start_catalog_refresh(credentials: dict[str, Any]) -> dict[str, str]:
    job_id = secrets.token_hex(12)
    set_job(job_id, status="queued", percent=0, message="Очікування запуску...")
    thread = threading.Thread(
        target=refresh_catalog_job,
        args=(job_id, credentials),
        daemon=True,
    )
    thread.start()
    return {"job_id": job_id}


def preview_job(
    job_id: str,
    mode: str,
    credentials: dict[str, Any],
    requested_articles: list[XmlRequest] | None = None,
    fresh_catalog: bool = False,
) -> None:
    base_settings = HOROSHOP_SETTINGS
    if base_settings is None:
        set_job(job_id, status="error", percent=100, message="Налаштування Хорошопа не ініціалізовано.")
        return
    if not SYNC_LOCK.acquire(blocking=False):
        set_job(job_id, status="error", percent=100, message="Інша операція вже виконується.")
        return
    try:
        settings = with_runtime_credentials(base_settings, credentials)
        set_job(job_id, status="running", percent=5, message="Побудова XML і плану...")
        build_xml(XML_CONFIG)
        xml_products = load_xml_products(XML_CONFIG["output_xml"])
        counts = image_counts(xml_products)
        target_articles = resolve_target_articles(
            mode=mode,
            xml_products=xml_products,
            requested_articles=requested_articles,
        )
        if not target_articles:
            set_job(job_id, status="done", percent=100, message="Немає артикулів для перевірки.", report={"preview": []})
            return
        client = HoroshopClient(settings)

        def export_progress(count: int, message: str) -> None:
            set_job(job_id, status="running", percent=25, message=message, exported=count)

        catalog, catalog_meta = load_or_refresh_catalog(
            client=client,
            force_refresh=fresh_catalog,
            progress=export_progress,
        )
        plan = prepare_import_plan(
            settings=settings,
            catalog=catalog,
            xml_products=xml_products,
            target_articles=target_articles,
        )

        def validation_progress(count: int, message: str) -> None:
            set_job(job_id, status="running", percent=70, message=message, checked_urls=count)

        valid, failed, validation_meta = validate_prepared_images(
            plan["prepared"],
            progress=validation_progress,
        )
        queue_meta: dict[str, Any] = {}
        if mode == "excel":
            queue_meta = queue_preview_articles(
                plan=plan,
                validation_failed=failed,
                counts=counts,
                event_type="excel",
            )
        report = {
            "prepared": len(valid),
            "skipped": plan["skipped"],
            "failed": failed,
            "preview": [] if mode == "excel" else plan["preview"],
            "checked_urls": validation_meta.get("checked_urls", 0),
            "catalog_source": catalog_meta.get("source"),
            "catalog_updated_at": catalog_meta.get("updated_at"),
            **queue_meta,
        }
        if mode == "excel":
            message = (
                f"Перевірка Excel готова. Додано в чергу: {queue_meta.get('queued', 0)}; "
                f"готово до оновлення: {len(valid)}; пропущено: {len(plan['skipped'])}; "
                f"помилок URL: {len(failed)}."
            )
        else:
            message = (
                f"Перевірка готова. До оновлення: {len(valid)}; "
                f"пропущено: {len(plan['skipped'])}; помилок URL: {len(failed)}."
            )
        set_job(
            job_id,
            status="done" if not failed else "warning",
            percent=100,
            message=message,
            report=report,
        )
    except (HoroshopError, OSError, ValueError) as error:
        set_job(job_id, status="error", percent=100, message=str(error))
    finally:
        SYNC_LOCK.release()


def start_preview(
    mode: str,
    credentials: dict[str, Any],
    articles: list[XmlRequest] | None = None,
    fresh_catalog: bool = False,
) -> dict[str, str]:
    job_id = secrets.token_hex(12)
    set_job(job_id, status="queued", percent=0, message="Очікування запуску перевірки...")
    thread = threading.Thread(
        target=preview_job,
        args=(job_id, mode, credentials, articles, fresh_catalog),
        daemon=True,
    )
    thread.start()
    return {"job_id": job_id}


def state_snapshot() -> dict[str, Any]:
    xml_meta: dict[str, Any] = {}
    output_xml = XML_CONFIG.get("output_xml")
    if isinstance(output_xml, Path) and output_xml.exists():
        xml_meta = {
            "path": str(output_xml),
            "updated_at": time.strftime(
                "%Y-%m-%d %H:%M:%S",
                time.localtime(output_xml.stat().st_mtime),
            ),
            "bytes": output_xml.stat().st_size,
        }
    with STATE_LOCK:
        state = get_state()
        dirty = sorted(state.dirty.values(), key=lambda item: str(item.get("article", "")).casefold())
        history = list(reversed(state.history[-50:]))
        archive = list(reversed(state.archive[-100:]))
        catalog_meta = state.catalog_meta()
    catalog_meta["age_seconds"] = catalog_age_seconds(
        str(catalog_meta.get("updated_at", ""))
    )
    return {
        "busy": SYNC_LOCK.locked(),
        "dirty_count": len(dirty),
        "dirty": dirty,
        "history": history,
        "archive": archive,
        "catalog": catalog_meta,
        "xml": xml_meta,
        "image_field": HOROSHOP_SETTINGS.image_field if HOROSHOP_SETTINGS else "",
        "import_mode": "replace" if HOROSHOP_SETTINGS and HOROSHOP_SETTINGS.override else "append",
        "remove_all_when_no_local_images": (
            HOROSHOP_SETTINGS.remove_all_when_no_local_images
            if HOROSHOP_SETTINGS
            else False
        ),
    }


def render_page() -> str:
    return f"""<!doctype html>
<html lang="uk">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>Python Scan - Horoshop Sync</title>
  <style>
    :root {{
      color-scheme: light;
      --bg: #f6f7f9;
      --ink: #17202a;
      --muted: #637083;
      --line: #cbd3df;
      --accent: #166534;
      --accent-strong: #14532d;
      --accent-soft: #dcfce7;
      --focus: #2563eb;
      --warn: #a16207;
      --bad: #b91c1c;
      --bad-soft: #fee2e2;
      --panel: #ffffff;
    }}
    * {{ box-sizing: border-box; }}
    body {{
      margin: 0;
      font-family: Arial, sans-serif;
      background: var(--bg);
      color: var(--ink);
    }}
    header {{
      display: flex;
      align-items: center;
      justify-content: space-between;
      gap: 16px;
      padding: 18px 24px;
      border-bottom: 1px solid var(--line);
      background: #fff;
    }}
    h1 {{ margin: 0; font-size: 22px; letter-spacing: 0; }}
    main {{ padding: 20px 24px 36px; display: grid; gap: 18px; }}
    .toolbar {{ display: flex; flex-wrap: wrap; gap: 10px; align-items: center; }}
    .panel {{
      display: grid;
      gap: 12px;
      padding: 14px;
      border: 1px solid var(--line);
      border-radius: 8px;
      background: var(--panel);
    }}
    .panel-title {{ margin: 0; font-size: 17px; letter-spacing: 0; }}
    .panel-description {{ margin: 2px 0 0; color: var(--muted); font-size: 14px; }}
    .fields {{ display: grid; grid-template-columns: minmax(180px, 1fr) minmax(180px, 1fr); gap: 10px; }}
    .excel-block {{ display: grid; gap: 10px; }}
    .file-action {{ display: grid; grid-template-columns: minmax(280px, 1fr) auto auto; gap: 10px; align-items: stretch; }}
    .template-link {{
      display: inline-flex;
      align-items: center;
      justify-content: center;
      min-height: 40px;
      border: 1px solid #94a3b8;
      border-radius: 6px;
      padding: 9px 12px;
      background: #fff;
      color: #334155;
      font-weight: 700;
      text-decoration: none;
    }}
    .template-link:hover {{ border-color: var(--focus); color: #1d4ed8; background: #eff6ff; }}
    .column-guide {{ display: flex; flex-wrap: wrap; gap: 8px; }}
    .column-guide span {{
      border: 1px solid #bfdbfe;
      border-radius: 6px;
      padding: 7px 9px;
      background: #eff6ff;
      color: #1e3a8a;
      font-size: 13px;
      font-weight: 700;
    }}
    .excel-dropzone {{
      position: relative;
      display: grid;
      gap: 5px;
      align-content: center;
      min-height: 86px;
      width: 100%;
      border: 2px solid var(--focus);
      border-radius: 8px;
      padding: 14px 46px 14px 14px;
      background: #eff6ff;
      color: #1e3a8a;
      cursor: pointer;
      font-weight: 800;
      transition: border-color .15s ease, box-shadow .15s ease, background-color .15s ease, color .15s ease;
    }}
    .excel-dropzone:hover,
    .excel-dropzone:focus-visible {{
      box-shadow: 0 0 0 3px rgba(37, 99, 235, .18);
      outline: 0;
    }}
    .excel-dropzone.has-file {{
      border-color: var(--accent);
      background: var(--accent-soft);
      color: var(--accent-strong);
    }}
    .excel-dropzone.is-empty {{
      border-color: #dc2626;
      background: #fff1f2;
      color: #991b1b;
    }}
    .excel-dropzone small {{ color: inherit; font-weight: 700; opacity: .82; }}
    .excel-input {{ position: absolute; inline-size: 1px; block-size: 1px; opacity: 0; pointer-events: none; }}
    .detach-excel {{
      position: absolute;
      top: 8px;
      right: 8px;
      display: none;
      width: 30px;
      height: 30px;
      padding: 0;
      border-radius: 999px;
      background: var(--accent-strong);
      color: #fff;
      font-size: 20px;
      line-height: 1;
    }}
    .excel-dropzone.has-file .detach-excel {{ display: inline-grid; place-items: center; }}
    .detach-excel:hover {{ background: var(--bad); }}
    .manual-action {{ display: grid; grid-template-columns: minmax(180px, 1fr) minmax(180px, 1fr) auto; gap: 10px; align-items: end; }}
    .manual-hint {{ grid-column: 1 / -1; }}
    label {{ display: grid; gap: 6px; font-weight: 700; color: #334155; }}
    input {{
      width: 100%;
      border: 1px solid var(--line);
      border-radius: 6px;
      padding: 10px 11px;
      font: inherit;
      color: var(--ink);
      background: #fff;
      transition: border-color .15s ease, box-shadow .15s ease, background-color .15s ease;
    }}
    input:hover {{ border-color: #94a3b8; }}
    input:focus {{
      border-color: var(--focus);
      box-shadow: 0 0 0 3px rgba(37, 99, 235, .16);
      outline: 0;
    }}
    input.input-error {{
      border-color: var(--bad);
      background: #fff7f7;
      box-shadow: 0 0 0 3px rgba(185, 28, 28, .12);
    }}
    button {{
      border: 0;
      border-radius: 6px;
      padding: 10px 14px;
      background: var(--accent);
      color: #fff;
      font-weight: 700;
      cursor: pointer;
      transition: transform .08s ease, background-color .15s ease, box-shadow .15s ease;
    }}
    button:hover {{ background: var(--accent-strong); }}
    button:active {{ transform: translateY(1px); }}
    button:focus-visible {{
      outline: 3px solid rgba(37, 99, 235, .28);
      outline-offset: 2px;
    }}
    button.secondary {{ background: #334155; }}
    button.danger {{ background: var(--bad); }}
    button.small {{ padding: 7px 10px; font-size: 13px; }}
    button:disabled {{ background: #94a3b8; cursor: wait; }}
    details.danger-zone {{
      border: 1px solid #fecaca;
      border-radius: 8px;
      background: #fff7f7;
      padding: 12px 14px;
    }}
    details.danger-zone summary {{
      cursor: pointer;
      font-weight: 800;
      color: var(--bad);
    }}
    .status {{
      padding: 12px 14px;
      border: 1px solid var(--line);
      border-radius: 8px;
      background: var(--panel);
      white-space: pre-wrap;
    }}
    .grid {{ display: grid; grid-template-columns: repeat(4, minmax(140px, 1fr)); gap: 10px; }}
    .metric {{ background: var(--panel); border: 1px solid var(--line); border-radius: 8px; padding: 12px; }}
    .metric span {{ display: block; color: var(--muted); font-size: 13px; }}
    .metric strong {{ display: block; margin-top: 6px; font-size: 19px; }}
    .table-wrap {{ overflow: auto; border: 1px solid var(--line); border-radius: 8px; background: var(--panel); }}
    table {{ width: 100%; border-collapse: collapse; min-width: 860px; }}
    th, td {{ padding: 10px 12px; border-bottom: 1px solid #e2e8f0; text-align: left; vertical-align: top; }}
    th {{ background: #eef2f7; font-size: 13px; color: #334155; }}
    tr:last-child td {{ border-bottom: 0; }}
    .muted {{ color: var(--muted); }}
    .status-dirty {{ color: var(--warn); font-weight: 700; }}
    .status-error {{ color: var(--bad); font-weight: 700; }}
    .status-skipped {{ color: #475569; font-weight: 700; }}
    .cache-fresh {{ color: var(--accent); font-weight: 700; }}
    .cache-stale {{ color: var(--warn); font-weight: 700; }}
    .cache-old {{ color: var(--bad); font-weight: 700; }}
    .report-panel {{ display: none; }}
    .report-panel.is-visible {{ display: block; }}
    .report-stats {{ display: flex; flex-wrap: wrap; gap: 8px; margin-bottom: 10px; }}
    .report-chip {{ border: 1px solid var(--line); border-radius: 8px; padding: 8px 10px; background: #fff; font-weight: 700; }}
    .inline-meta {{ display: grid; gap: 4px; padding: 10px 12px; border: 1px solid var(--line); border-radius: 8px; background: #fff; }}
    .notice {{ padding: 11px 12px; border-radius: 8px; border: 1px solid var(--line); background: #fff; font-weight: 700; }}
    .notice-warning {{ border-color: #fbbf24; background: #fffbeb; color: #78350f; }}
    .notice-success {{ border-color: #86efac; background: var(--accent-soft); color: var(--accent-strong); }}
    .is-hidden {{ display: none; }}
    .section-heading {{ display: flex; align-items: center; justify-content: space-between; gap: 10px; flex-wrap: wrap; }}
    .empty {{ padding: 18px; color: var(--muted); }}
    @media (max-width: 760px) {{
      header {{ align-items: flex-start; flex-direction: column; }}
      main {{ padding: 16px; }}
      .grid {{ grid-template-columns: 1fr; }}
      .fields {{ grid-template-columns: 1fr; }}
      .file-action {{ grid-template-columns: 1fr; }}
      .manual-action {{ grid-template-columns: 1fr; }}
      button {{ width: 100%; }}
    }}
  </style>
</head>
<body>
  <header>
    <div>
      <h1>Horoshop Sync</h1>
      <div class="muted">Локальна панель без окремого пароля</div>
    </div>
    <div class="toolbar">
      <button id="refreshButton" class="secondary" type="button">Оновити статус</button>
      <button id="rebuildButton" class="secondary" type="button">Перебудувати XML</button>
      <button id="previewDirtyButton" class="secondary" type="button">Перевірити змінені</button>
      <button id="syncDirtyButton" type="button">Оновити змінені артикули</button>
    </div>
  </header>
  <main>
    <section class="panel">
      <div>
        <h2 class="panel-title">Доступ до Хорошоп</h2>
        <p class="panel-description">Дані потрібні лише під час перевірки або оновлення.</p>
      </div>
      <div class="fields">
        <label>Логін Хорошоп
          <input id="shopLogin" autocomplete="username" placeholder="api@example.com">
        </label>
        <label>Пароль Хорошоп
          <input id="shopPassword" type="password" autocomplete="current-password">
        </label>
      </div>
      <div id="credentialNotice" class="notice notice-warning is-hidden">
        Введіть логін і пароль Хорошопа перед перевіркою, експортом або оновленням.
      </div>
    </section>
    <section class="panel">
      <div class="section-heading">
        <div>
          <h2 class="panel-title">Excel-імпорт</h2>
          <p class="panel-description">Додайте потрібні артикули до черги або оновіть їх одразу.</p>
        </div>
        <a class="template-link" href="/api/excel-template" download="horoshop_import_template.xlsx">Завантажити шаблон</a>
      </div>
      <div class="column-guide">
        <span>1. Артикул</span>
        <span>2. Бренд (необов'язково)</span>
      </div>
      <div class="excel-block">
        <div class="file-action">
          <div id="excelDropzone" class="excel-dropzone is-empty" role="button" tabindex="0" aria-label="Прикріпити Excel зі списком артикулів">
            <span id="excelDropTitle">Прикріпити Excel зі списком артикулів</span>
            <small id="excelDropHint">Файл не вибрано. Натисніть тут, щоб вибрати .xlsx або .xlsm.</small>
            <button id="detachExcelButton" class="detach-excel" type="button" aria-label="Відкріпити Excel">×</button>
            <input id="excelFile" class="excel-input" type="file" accept=".xlsx,.xlsm">
          </div>
          <button id="previewExcelButton" class="secondary" type="button">Перевірити Excel і додати в чергу</button>
          <button id="syncExcelButton" type="button">Оновити артикули з Excel</button>
        </div>
      </div>
      <div class="muted">Порожній бренд не обробляється як brand=&quot;&quot;: буде використано перший варіант артикула з XML.</div>
    </section>
    <section class="panel">
      <div>
        <h2 class="panel-title">Ручне додавання</h2>
        <p class="panel-description">Для одиничного товару або швидкої перевірки.</p>
      </div>
      <div class="manual-action">
        <label>Додати артикул у чергу вручну
          <input id="manualArticle" autocomplete="off" placeholder="Наприклад: X33">
        </label>
        <label>Бренд (необов'язково)
          <input id="manualBrand" autocomplete="off" placeholder="Наприклад: BRUDER">
        </label>
        <button id="manualAddButton" class="secondary" type="button">Перевірити і додати</button>
        <div id="manualAddHint" class="muted manual-hint">Порожній бренд не фільтрує: буде використано перший варіант артикула з XML.</div>
      </div>
    </section>
    <section class="panel">
      <div class="toolbar">
        <label style="display:flex;align-items:center;gap:8px;font-weight:700;">
          <input id="freshCatalog" type="checkbox" style="width:auto;">
          Зробити свіжий експорт каталогу перед запуском
        </label>
        <button id="refreshCatalogButton" class="secondary" type="button">Згенерувати експорт сайту</button>
      </div>
      <div class="inline-meta">
        <div id="catalogStatus" class="muted">Експорт сайту ще не завантажено.</div>
        <div class="muted">Останнє оновлення: <strong id="catalogUpdatedAt">-</strong></div>
        <div class="muted">Локальний файл: <span id="catalogLocalPath">-</span></div>
      </div>
      <div class="muted">Для заповненого бренду спочатку шукається точний збіг, потім збіг без урахування регістру; якщо варіанту немає, буде показано доступні бренди.</div>
    </section>
    <div class="grid">
      <div class="metric"><span>Змінені артикули</span><strong id="dirtyCount">0</strong></div>
      <div class="metric"><span>Стан синхронізації</span><strong id="busyState">-</strong></div>
      <div class="metric"><span>Галерея</span><strong id="imageField">-</strong></div>
      <div class="metric"><span>Режим</span><strong id="importMode">-</strong></div>
    </div>
    <div id="status" class="status">Завантаження...</div>
    <section id="reportPanel" class="panel report-panel">
      <h2>Результат останньої операції</h2>
      <div id="reportContent"></div>
    </section>
    <details class="danger-zone">
      <summary>Небезпечні дії</summary>
      <p class="muted">Повне оновлення проходить по всіх артикулах XML. Для контрольованої роботи краще використовувати Excel-список.</p>
      <button id="syncAllButton" class="danger" type="button">Повне оновлення</button>
    </details>
    <section>
      <div class="section-heading">
        <h2>Черга змін</h2>
        <div class="toolbar">
          <button id="syncQueueButton" type="button">Оновити всі</button>
          <button id="skipQueueButton" class="secondary" type="button">Пропустити всі</button>
          <button id="deleteQueueButton" class="danger" type="button">Видалити чергу</button>
        </div>
      </div>
      <div id="dirtyTable" class="table-wrap"><div class="empty">Завантаження...</div></div>
    </section>
    <section>
      <div class="section-heading">
        <h2>Останні події</h2>
        <button id="archiveHistoryButton" class="secondary" type="button">Архівувати події</button>
      </div>
      <div class="toolbar">
        <button id="archiveHistoryInlineButton" class="secondary" type="button">Перенести останні події в архів</button>
      </div>
      <div id="historyTable" class="table-wrap"><div class="empty">Завантаження...</div></div>
      <details>
        <summary>Архів</summary>
        <div id="archiveTable" class="table-wrap"><div class="empty">Архів порожній.</div></div>
      </details>
    </section>
  </main>
  <script>
    const statusBox = document.getElementById('status');
    const dirtyCount = document.getElementById('dirtyCount');
    const busyState = document.getElementById('busyState');
    const imageField = document.getElementById('imageField');
    const importMode = document.getElementById('importMode');
    const dirtyTable = document.getElementById('dirtyTable');
    const historyTable = document.getElementById('historyTable');
    const reportPanel = document.getElementById('reportPanel');
    const reportContent = document.getElementById('reportContent');
    const shopLogin = document.getElementById('shopLogin');
    const shopPassword = document.getElementById('shopPassword');
    const credentialNotice = document.getElementById('credentialNotice');
    const manualArticle = document.getElementById('manualArticle');
    const manualBrand = document.getElementById('manualBrand');
    const manualAddButton = document.getElementById('manualAddButton');
    const manualAddHint = document.getElementById('manualAddHint');
    const excelFile = document.getElementById('excelFile');
    const excelDropzone = document.getElementById('excelDropzone');
    const excelDropTitle = document.getElementById('excelDropTitle');
    const excelDropHint = document.getElementById('excelDropHint');
    const freshCatalog = document.getElementById('freshCatalog');
    const catalogStatus = document.getElementById('catalogStatus');
    const catalogUpdatedAt = document.getElementById('catalogUpdatedAt');
    const catalogLocalPath = document.getElementById('catalogLocalPath');
    const archiveTable = document.getElementById('archiveTable');
    const buttons = [
      'refreshButton', 'rebuildButton', 'syncDirtyButton', 'syncAllButton',
      'previewDirtyButton', 'refreshCatalogButton', 'previewExcelButton',
      'syncExcelButton', 'detachExcelButton', 'manualAddButton', 'syncQueueButton',
      'skipQueueButton', 'deleteQueueButton', 'archiveHistoryButton',
      'archiveHistoryInlineButton'
    ].map((id) => document.getElementById(id));
    let activeJob = '';

    function escapeHtml(value) {{
      return String(value ?? '')
        .replaceAll('&', '&amp;')
        .replaceAll('<', '&lt;')
        .replaceAll('>', '&gt;')
        .replaceAll('"', '&quot;')
        .replaceAll("'", '&#39;');
    }}

    function setBusy(isBusy) {{
      buttons.forEach((button) => {{
        if (button.id !== 'refreshButton') button.disabled = isBusy;
      }});
    }}

    function updateExcelDropzone() {{
      const hasFile = Boolean(excelFile.files.length);
      excelDropzone.classList.toggle('has-file', hasFile);
      excelDropzone.classList.toggle('is-empty', !hasFile);
      if (hasFile) {{
        const file = excelFile.files[0];
        excelDropTitle.textContent = file.name;
        excelDropHint.textContent = 'Excel прикріплено. Натисніть ×, щоб відкріпити, або натисніть область, щоб замінити файл.';
      }} else {{
        excelDropTitle.textContent = 'Прикріпити Excel зі списком артикулів';
        excelDropHint.textContent = 'Файл не вибрано. Натисніть тут, щоб вибрати .xlsx або .xlsm.';
      }}
    }}

    function requireExcelFile() {{
      if (excelFile.files.length) return;
      updateExcelDropzone();
      excelDropzone.focus();
      throw new Error('Виберіть Excel-файл зі списком артикулів.');
    }}

    async function api(path, options = {{}}) {{
      const response = await fetch(path, {{
        cache: 'no-store',
        headers: {{ 'Accept': 'application/json' }},
        ...options
      }});
      if (!response.ok) {{
        let message = response.statusText;
        try {{
          const payload = await response.json();
          message = payload.detail || message;
        }} catch (_) {{}}
        throw new Error(message);
      }}
      return response.json();
    }}

    function credentialFormData() {{
      const login = shopLogin.value.trim();
      const password = shopPassword.value;
      if (!login || !password) {{
        showCredentialNotice('Введіть логін і пароль Хорошопа. Без них сайт не дозволить перевірити каталог або оновити фото.');
        throw new Error('Введіть логін і пароль Хорошопа.');
      }}
      clearCredentialNotice();
      const data = new FormData();
      data.append('login', login);
      data.append('password', password);
      data.append('fresh_catalog', freshCatalog.checked ? '1' : '0');
      return data;
    }}

    function showCredentialNotice(message) {{
      credentialNotice.textContent = message;
      credentialNotice.classList.remove('is-hidden');
      credentialNotice.classList.add('notice-warning');
      shopLogin.classList.toggle('input-error', !shopLogin.value.trim());
      shopPassword.classList.toggle('input-error', !shopPassword.value);
      if (!shopLogin.value.trim()) {{
        shopLogin.focus();
      }} else if (!shopPassword.value) {{
        shopPassword.focus();
      }}
    }}

    function clearCredentialNotice() {{
      if (shopLogin.value.trim() && shopPassword.value) {{
        credentialNotice.classList.add('is-hidden');
        shopLogin.classList.remove('input-error');
        shopPassword.classList.remove('input-error');
      }}
    }}

    function requireFullUpdateCredentials() {{
      const login = shopLogin.value.trim();
      const password = shopPassword.value;
      if (!login || !password) {{
        showCredentialNotice('Для повного оновлення обов’язково введіть логін і пароль Хорошопа, а потім повторіть їх у підтвердженні.');
        throw new Error('Введіть логін і пароль Хорошопа.');
      }}
      clearCredentialNotice();
      const repeatedLogin = window.prompt('Повторно введіть логін Хорошопа для повного оновлення:') || '';
      if (repeatedLogin.trim() !== login) {{
        throw new Error('Повторний логін не збігається. Повне оновлення скасовано.');
      }}
      const repeatedPassword = window.prompt('Повторно введіть пароль Хорошопа для повного оновлення:') || '';
      if (repeatedPassword !== password) {{
        throw new Error('Повторний пароль не збігається. Повне оновлення скасовано.');
      }}
    }}

    function renderTable(target, rows, mode = 'queue') {{
      if (!rows.length) {{
        target.innerHTML = '<div class="empty">Немає записів.</div>';
        return;
      }}
      let markup = '<table><thead><tr>' +
        '<th>Артикул</th><th>Бренд</th><th>Статус</th><th>Фото</th><th>Оновлено</th><th>Повідомлення</th><th>Дія</th>' +
        '</tr></thead><tbody>';
      for (const row of rows) {{
        const status = escapeHtml(row.status || '');
        const article = escapeHtml(row.article || '');
        const brand = escapeHtml(row.brand || '');
        const updatedAt = escapeHtml(row.updated_at || row.first_seen_at || '');
        let actions = '<span class="muted">-</span>';
        if (mode === 'queue') {{
          actions = '<button class="small article-sync" type="button" data-article="' + article + '" data-brand="' + brand + '">Оновити</button>' +
            '<button class="small secondary article-skip" type="button" data-article="' + article + '" data-brand="' + brand + '">Пропустити</button>';
        }} else if (mode === 'history') {{
          actions = '<button class="small secondary history-archive" type="button" data-article="' + article + '" data-brand="' + brand + '" data-updated="' + updatedAt + '">Архівувати</button>';
        }} else if (mode === 'archive') {{
          actions = '<span class="muted">В архіві</span>';
        }}
        markup += '<tr>' +
          '<td><strong>' + article + '</strong>' +
          (row.remote_article ? '<div class="muted">H: ' + escapeHtml(row.remote_article) + '</div>' : '') +
          '</td>' +
          '<td>' + (brand || '<span class="muted">-</span>') + '</td>' +
          '<td class="status-' + status + '">' + status + '</td>' +
          '<td>' + escapeHtml(row.image_count ?? '') + '</td>' +
          '<td>' + updatedAt + '</td>' +
          '<td>' + escapeHtml(row.message || '') + '</td>' +
          '<td><div class="toolbar">' +
          actions +
          '</div></td>' +
          '</tr>';
      }}
      markup += '</tbody></table>';
      target.innerHTML = markup;
      target.querySelectorAll('.article-sync').forEach((button) => {{
        button.addEventListener('click', async () => {{
          try {{
            const data = credentialFormData();
            const article = button.getAttribute('data-article') || '';
            data.append('brand', button.getAttribute('data-brand') || '');
            const result = await api('/api/sync/article/' + encodeURIComponent(article), {{
              method: 'POST',
              body: data
            }});
            pollJob(result.job_id);
          }} catch (error) {{
            statusBox.textContent = error.message;
          }}
        }});
      }});
      target.querySelectorAll('.article-skip').forEach((button) => {{
        button.addEventListener('click', async () => {{
          try {{
            const article = button.getAttribute('data-article') || '';
            const data = new FormData();
            data.append('brand', button.getAttribute('data-brand') || '');
            await api('/api/dirty/' + encodeURIComponent(article) + '/skip', {{
              method: 'POST',
              body: data
            }});
            await refreshState();
          }} catch (error) {{
            statusBox.textContent = error.message;
          }}
        }});
      }});
      target.querySelectorAll('.history-archive').forEach((button) => {{
        button.addEventListener('click', async () => {{
          try {{
            const data = new FormData();
            data.append('article', button.getAttribute('data-article') || '');
            data.append('brand', button.getAttribute('data-brand') || '');
            data.append('updated_at', button.getAttribute('data-updated') || '');
            await api('/api/history/archive-item', {{
              method: 'POST',
              body: data
            }});
            await refreshState();
          }} catch (error) {{
            statusBox.textContent = error.message;
          }}
        }});
      }});
    }}

    function renderOperationReport(report) {{
      if (!report) {{
        reportPanel.classList.remove('is-visible');
        reportContent.replaceChildren();
        return;
      }}
      const failed = Array.isArray(report.failed) ? report.failed : [];
      const skipped = Array.isArray(report.skipped) ? report.skipped : [];
      const preview = Array.isArray(report.preview) ? report.preview : [];
      const chips = [
        ['Готово до оновлення', report.prepared ?? 0],
        ['Успішно оновлено', report.imported ?? 0],
        ['Пропущено', skipped.length],
        ['Помилки', failed.length],
        ['Перевірено URL', report.checked_urls ?? 0]
      ];
      if (report.queued !== undefined) {{
        chips.splice(1, 0, ['Додано в чергу', report.queued ?? 0]);
      }}
      let markup = '<div class="report-stats">';
      for (const [label, value] of chips) {{
        markup += '<div class="report-chip">' + escapeHtml(label) + ': ' + escapeHtml(value) + '</div>';
      }}
      markup += '</div>';
      if (failed.length) {{
        markup += '<h3>Помилки</h3><ul>';
        failed.slice(0, 100).forEach((item) => {{
          markup += '<li><strong>' + escapeHtml(item.article || '') + '</strong>: ' + escapeHtml(item.message || '') + '</li>';
        }});
        markup += '</ul>';
      }}
      if (skipped.length) {{
        markup += '<h3>Пропущено</h3><ul>';
        skipped.slice(0, 100).forEach((item) => {{
          markup += '<li><strong>' + escapeHtml(item.article || '') + '</strong>: ' + escapeHtml(item.message || '') + '</li>';
        }});
        markup += '</ul>';
      }}
      if (preview.length) {{
        markup += '<h3>План</h3><div class="table-wrap"><table><thead><tr>' +
          '<th>Артикул</th><th>H-артикул</th><th>Статус</th><th>Фото</th><th>Приклади URL</th>' +
          '</tr></thead><tbody>';
        preview.slice(0, 150).forEach((item) => {{
          const samples = Array.isArray(item.sample_images) ? item.sample_images.slice(0, 3).join('\\n') : '';
          markup += '<tr><td><strong>' + escapeHtml(item.article || '') + '</strong></td>' +
            '<td>' + escapeHtml(item.remote_article || '') + '</td>' +
            '<td>' + escapeHtml(item.status || '') + '</td>' +
            '<td>' + escapeHtml(item.image_count ?? 0) + '</td>' +
            '<td style="white-space:pre-wrap;">' + escapeHtml(samples) + '</td></tr>';
        }});
        markup += '</tbody></table></div>';
      }}
      reportContent.innerHTML = markup;
      reportPanel.classList.add('is-visible');
    }}

    async function refreshState() {{
      const state = await api('/api/state');
      dirtyCount.textContent = state.dirty_count;
      busyState.textContent = state.busy ? 'Працює' : 'Вільно';
      imageField.textContent = state.image_field || '-';
      importMode.textContent = state.import_mode || '-';
      renderTable(dirtyTable, state.dirty || [], 'queue');
      renderTable(historyTable, state.history || [], 'history');
      renderTable(archiveTable, state.archive || [], 'archive');
      const xml = state.xml || {{}};
      const catalog = state.catalog || {{}};
      const age = Number(catalog.age_seconds);
      const cacheClass = !catalog.has_cache
        ? 'muted'
        : age > 7 * 24 * 3600
          ? 'cache-old'
          : age > 24 * 3600
            ? 'cache-stale'
            : 'cache-fresh';
      catalogStatus.className = cacheClass;
      catalogStatus.textContent = catalog.has_cache
        ? 'Експорт сайту отримано: ' + (catalog.products_count || 0) +
          ' товарів, оновлено ' + (catalog.updated_at || '-')
        : 'Експорт сайту ще не створено. Перед першим оновленням буде виконано свіжий експорт каталогу.';
      catalogUpdatedAt.textContent = catalog.updated_at || '-';
      catalogLocalPath.textContent = catalog.local_path || '-';
      statusBox.textContent =
        'XML: ' + (xml.path || '-') + '\\n' +
        'Оновлено: ' + (xml.updated_at || '-') + '\\n' +
        'Розмір: ' + (xml.bytes || 0) + ' байт';
      setBusy(Boolean(state.busy || activeJob));
    }}

    async function pollJob(jobId) {{
      activeJob = jobId;
      setBusy(true);
      while (activeJob === jobId) {{
        const job = await api('/api/progress/' + encodeURIComponent(jobId));
        statusBox.textContent = Math.round(job.percent || 0) + '% - ' + (job.message || '');
        if (job.status === 'done' || job.status === 'warning' || job.status === 'error') {{
          activeJob = '';
          renderOperationReport(job.report);
          await refreshState();
          return;
        }}
        await new Promise((resolve) => setTimeout(resolve, 1500));
      }}
    }}

    document.getElementById('refreshButton').addEventListener('click', refreshState);
    document.getElementById('rebuildButton').addEventListener('click', async () => {{
      setBusy(true);
      try {{
        const result = await api('/api/rebuild', {{ method: 'POST' }});
        statusBox.textContent = result.message;
        await refreshState();
      }} catch (error) {{
        statusBox.textContent = error.message;
      }} finally {{
        setBusy(false);
      }}
    }});
    document.getElementById('previewDirtyButton').addEventListener('click', async () => {{
      try {{
        const result = await api('/api/preview/dirty', {{ method: 'POST', body: credentialFormData() }});
        pollJob(result.job_id);
      }} catch (error) {{
        statusBox.textContent = error.message;
      }}
    }});
    async function syncDirtyQueue() {{
      try {{
        const result = await api('/api/sync/dirty', {{ method: 'POST', body: credentialFormData() }});
        pollJob(result.job_id);
      }} catch (error) {{
        statusBox.textContent = error.message;
      }}
    }}

    document.getElementById('syncDirtyButton').addEventListener('click', syncDirtyQueue);
    document.getElementById('syncQueueButton').addEventListener('click', syncDirtyQueue);
    document.getElementById('syncAllButton').addEventListener('click', async () => {{
      const warning = [
        'Повне оновлення пройде по всіх артикулах із XML.',
        'У режимі replace галереї товарів будуть очищені й завантажені заново.',
        'Якщо в XML неповний набір фото, на сайті залишиться саме неповний набір.',
        'Для контрольованого оновлення краще використати Excel-файл зі списком потрібних артикулів.'
      ].join('\\n');
      if (!window.confirm(warning)) return;
      const typed = window.prompt('Для запуску введіть: ПОВНЕ ОНОВЛЕННЯ');
      if (typed !== 'ПОВНЕ ОНОВЛЕННЯ') {{
        statusBox.textContent = 'Повне оновлення скасовано.';
        return;
      }}
      try {{
        requireFullUpdateCredentials();
        const result = await api('/api/sync/all', {{ method: 'POST', body: credentialFormData() }});
        pollJob(result.job_id);
      }} catch (error) {{
        statusBox.textContent = error.message;
      }}
    }});
    document.getElementById('previewExcelButton').addEventListener('click', async () => {{
      try {{
        requireExcelFile();
        const data = credentialFormData();
        data.append('file', excelFile.files[0], excelFile.files[0].name);
        const result = await api('/api/preview/excel', {{
          method: 'POST',
          body: data
        }});
        pollJob(result.job_id);
      }} catch (error) {{
        statusBox.textContent = error.message;
      }}
    }});
    document.getElementById('syncExcelButton').addEventListener('click', async () => {{
      try {{
        requireExcelFile();
        const data = credentialFormData();
        data.append('file', excelFile.files[0], excelFile.files[0].name);
        const result = await api('/api/sync/excel', {{
          method: 'POST',
          body: data
        }});
        pollJob(result.job_id);
      }} catch (error) {{
        statusBox.textContent = error.message;
      }}
    }});
    excelDropzone.addEventListener('click', () => {{
      excelFile.click();
    }});
    excelDropzone.addEventListener('keydown', (event) => {{
      if (event.key === 'Enter' || event.key === ' ') {{
        event.preventDefault();
        excelFile.click();
      }}
    }});
    excelFile.addEventListener('change', () => {{
      updateExcelDropzone();
      if (excelFile.files.length) {{
        statusBox.textContent = 'Excel-файл прикріплено: ' + excelFile.files[0].name;
      }}
    }});
    document.getElementById('detachExcelButton').addEventListener('click', (event) => {{
      event.stopPropagation();
      if (!excelFile.files.length) {{
        statusBox.textContent = 'Excel-файл не прикріплено.';
        updateExcelDropzone();
        return;
      }}
      const filename = excelFile.files[0].name;
      excelFile.value = '';
      excelFile.classList.remove('input-error');
      updateExcelDropzone();
      statusBox.textContent = 'Excel-файл відкріплено: ' + filename + '. Черга змін не змінена.';
    }});
    document.getElementById('refreshCatalogButton').addEventListener('click', async () => {{
      try {{
        const data = credentialFormData();
        const result = await api('/api/catalog/refresh', {{
          method: 'POST',
          body: data
        }});
        pollJob(result.job_id);
      }} catch (error) {{
        statusBox.textContent = error.message;
      }}
    }});

    document.getElementById('skipQueueButton').addEventListener('click', async () => {{
      if (!window.confirm('Пропустити всі артикули з черги? Вони перейдуть в останні події зі статусом skipped.')) return;
      try {{
        const result = await api('/api/dirty/skip-all', {{ method: 'POST' }});
        statusBox.textContent = 'Пропущено артикулів: ' + (result.skipped || 0);
        await refreshState();
      }} catch (error) {{
        statusBox.textContent = error.message;
      }}
    }});

    document.getElementById('deleteQueueButton').addEventListener('click', async () => {{
      if (!window.confirm('Видалити всю чергу без сліду? Це корисно, якщо випадково вибрали не той Excel-файл.')) return;
      try {{
        await api('/api/clear-dirty', {{ method: 'POST' }});
        renderOperationReport(null);
        statusBox.textContent = 'Чергу видалено без запису в історію.';
        await refreshState();
      }} catch (error) {{
        statusBox.textContent = error.message;
      }}
    }});

    async function archiveHistory() {{
      try {{
        const result = await api('/api/history/archive', {{ method: 'POST' }});
        statusBox.textContent = 'Події перенесено в архів: ' + (result.archived || 0);
        await refreshState();
      }} catch (error) {{
        statusBox.textContent = error.message;
      }}
    }}
    document.getElementById('archiveHistoryButton').addEventListener('click', archiveHistory);
    document.getElementById('archiveHistoryInlineButton').addEventListener('click', archiveHistory);
    shopLogin.addEventListener('input', clearCredentialNotice);
    shopPassword.addEventListener('input', clearCredentialNotice);

    async function addManualArticle() {{
        const article = manualArticle.value.trim();
      const brand = manualBrand.value.trim();
      if (!article) {{
        manualArticle.classList.add('input-error');
        manualArticle.focus();
        manualAddHint.className = 'notice notice-warning';
        manualAddHint.textContent = 'Введіть артикул, наприклад X33, щоб перевірити фото і додати його в чергу.';
        return;
      }}
      manualArticle.classList.remove('input-error');
      manualAddHint.className = 'muted';
      manualAddHint.textContent = 'Перевіряю локальний XML і фото для артикула...';
      try {{
        const data = new FormData();
        data.append('article', article);
        data.append('brand', brand);
        const result = await api('/api/dirty/manual', {{
          method: 'POST',
          body: data
        }});
        manualArticle.value = '';
        manualBrand.value = '';
        manualAddHint.className = 'notice notice-success';
        manualAddHint.textContent = 'Артикул ' + result.article + (result.brand ? ' (' + result.brand + ')' : '') + ' додано в чергу. Фото: ' + result.image_count + '.';
        statusBox.textContent = manualAddHint.textContent;
        await refreshState();
      }} catch (error) {{
        manualArticle.classList.add('input-error');
        manualArticle.focus();
        manualAddHint.className = 'notice notice-warning';
        manualAddHint.textContent = error.message;
        statusBox.textContent = error.message;
      }}
    }}

    manualAddButton.addEventListener('click', addManualArticle);
    manualArticle.addEventListener('keydown', (event) => {{
      if (event.key === 'Enter') {{
        event.preventDefault();
        addManualArticle();
      }}
    }});
    manualArticle.addEventListener('input', () => {{
      manualArticle.classList.remove('input-error');
      if (!manualArticle.value.trim()) {{
        manualAddHint.className = 'muted';
        manualAddHint.textContent = 'Перевіряє, чи є фото в локальному XML, і додає артикул у чергу змін.';
      }}
    }});

    refreshState().catch((error) => {{
      statusBox.textContent = error.message;
    }});
    updateExcelDropzone();
    window.setInterval(() => {{
      if (!activeJob) refreshState().catch(() => {{}});
    }}, 10000);
  </script>
</body>
</html>"""


@asynccontextmanager
async def lifespan(_: FastAPI):
    global OBSERVER
    OBSERVER = start_observer()
    try:
        yield
    finally:
        if OBSERVER is not None:
            OBSERVER.stop()
            OBSERVER.join()
            OBSERVER = None


app = FastAPI(title="Python Scan Horoshop Sync", lifespan=lifespan)


@app.get("/", response_class=HTMLResponse)
def index(request: Request) -> str:
    return render_page()


@app.get("/status")
def public_status() -> dict[str, Any]:
    return {
        "ok": True,
        "busy": SYNC_LOCK.locked(),
        "dirty_count": len(get_state().dirty) if STATE is not None else 0,
        "server": {
            "port": SERVER_SETTINGS.get("port"),
            "mode": "horoshop_sync",
        },
    }


@app.get("/api/state")
def api_state(request: Request) -> dict[str, Any]:
    protected_json(request)
    return state_snapshot()


@app.get("/api/excel-template")
def api_excel_template() -> Response:
    return Response(
        content=build_excel_template(),
        media_type=(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        ),
        headers={
            "Content-Disposition": (
                'attachment; filename="horoshop_import_template.xlsx"'
            )
        },
    )


@app.post("/api/rebuild")
def api_rebuild(request: Request) -> dict[str, Any]:
    protected_json(request)
    product_count, image_count = build_xml(XML_CONFIG)
    return {
        "ok": True,
        "message": f"XML перебудовано: артикулів {product_count}, фото {image_count}.",
    }


@app.post("/api/sync/dirty")
async def api_sync_dirty(request: Request) -> dict[str, str]:
    protected_json(request)
    form = await request.form()
    return start_sync(
        "dirty",
        credentials_from_form(form),
        fresh_catalog=fresh_catalog_from_form(form),
    )


@app.post("/api/sync/all")
async def api_sync_all(request: Request) -> dict[str, str]:
    protected_json(request)
    form = await request.form()
    return start_sync(
        "all",
        credentials_from_form(form),
        fresh_catalog=fresh_catalog_from_form(form),
    )


@app.post("/api/sync/article/{article}")
async def api_sync_article(article: str, request: Request) -> dict[str, str]:
    protected_json(request)
    form = await request.form()
    return start_sync(
        "article",
        credentials_from_form(form),
        [XmlRequest(article, str(form.get("brand", "")))],
        fresh_catalog=fresh_catalog_from_form(form),
    )


@app.post("/api/sync/excel")
async def api_sync_excel(request: Request) -> dict[str, Any]:
    protected_json(request)
    form = await request.form()
    uploaded = form.get("file")
    if not isinstance(uploaded, StarletteUploadFile):
        raise HTTPException(status_code=400, detail="Excel-файл не передано.")
    filename = uploaded.filename or ""
    if Path(filename).suffix.lower() not in {".xlsx", ".xlsm"}:
        raise HTTPException(status_code=400, detail="Підтримуються лише .xlsx та .xlsm.")
    data = await uploaded.read()
    if len(data) > 20 * 1024 * 1024:
        raise HTTPException(status_code=413, detail="Excel-файл більший за 20 МБ.")
    try:
        articles = parse_excel_articles(data)
    except Exception as error:
        raise HTTPException(status_code=400, detail=f"Не вдалося прочитати Excel: {error}") from error
    if not articles:
        raise HTTPException(status_code=400, detail="В Excel не знайдено артикулів у першому стовпці.")
    result = start_sync(
        "excel",
        credentials_from_form(form),
        articles,
        fresh_catalog=fresh_catalog_from_form(form),
    )
    result["articles_count"] = len(articles)
    return result


@app.post("/api/preview/dirty")
async def api_preview_dirty(request: Request) -> dict[str, str]:
    protected_json(request)
    form = await request.form()
    return start_preview(
        "dirty",
        credentials_from_form(form),
        fresh_catalog=fresh_catalog_from_form(form),
    )


@app.post("/api/preview/excel")
async def api_preview_excel(request: Request) -> dict[str, Any]:
    protected_json(request)
    form = await request.form()
    uploaded = form.get("file")
    if not isinstance(uploaded, StarletteUploadFile):
        raise HTTPException(status_code=400, detail="Excel-файл не передано.")
    filename = uploaded.filename or ""
    if Path(filename).suffix.lower() not in {".xlsx", ".xlsm"}:
        raise HTTPException(status_code=400, detail="Підтримуються лише .xlsx та .xlsm.")
    data = await uploaded.read()
    if len(data) > 20 * 1024 * 1024:
        raise HTTPException(status_code=413, detail="Excel-файл більший за 20 МБ.")
    try:
        articles = parse_excel_articles(data)
    except Exception as error:
        raise HTTPException(status_code=400, detail=f"Не вдалося прочитати Excel: {error}") from error
    if not articles:
        raise HTTPException(status_code=400, detail="В Excel не знайдено артикулів у першому стовпці.")
    result = start_preview(
        "excel",
        credentials_from_form(form),
        articles,
        fresh_catalog=fresh_catalog_from_form(form),
    )
    result["articles_count"] = len(articles)
    return result


@app.post("/api/catalog/refresh")
async def api_catalog_refresh(request: Request) -> dict[str, str]:
    protected_json(request)
    form = await request.form()
    return start_catalog_refresh(credentials_from_form(form))


@app.get("/api/progress/{job_id}")
def api_progress(job_id: str, request: Request) -> dict[str, Any]:
    protected_json(request)
    return get_job(job_id)


@app.post("/api/dirty/manual")
async def api_mark_dirty_manual(request: Request) -> dict[str, Any]:
    protected_json(request)
    form = await request.form()
    try:
        return add_manual_dirty_article(
            str(form.get("article", "")),
            str(form.get("brand", "")),
        )
    except ValueError as error:
        raise HTTPException(status_code=400, detail=str(error)) from error


@app.post("/api/dirty/skip-all")
def api_skip_all_dirty(request: Request) -> dict[str, Any]:
    protected_json(request)
    with STATE_LOCK:
        state = get_state()
        skipped = state.skip_all_dirty()
        state.save()
    return {"ok": True, "skipped": skipped}


@app.post("/api/dirty/{article}")
async def api_mark_dirty(article: str, request: Request) -> dict[str, Any]:
    protected_json(request)
    form = await request.form()
    try:
        return add_manual_dirty_article(article, str(form.get("brand", "")))
    except ValueError as error:
        raise HTTPException(status_code=400, detail=str(error)) from error


@app.post("/api/dirty/{article}/skip")
async def api_skip_dirty(article: str, request: Request) -> dict[str, Any]:
    protected_json(request)
    form = await request.form()
    with STATE_LOCK:
        state = get_state()
        skipped = state.skip_dirty(article, str(form.get("brand", "")))
        state.save()
    if not skipped:
        raise HTTPException(status_code=404, detail="Артикул не знайдено в черзі.")
    return {"ok": True}


@app.post("/api/clear-dirty")
def api_clear_dirty(request: Request) -> dict[str, Any]:
    protected_json(request)
    with STATE_LOCK:
        state = get_state()
        state.clear_dirty()
        state.save()
    return {"ok": True}


@app.post("/api/history/archive")
def api_archive_history(request: Request) -> dict[str, Any]:
    protected_json(request)
    with STATE_LOCK:
        state = get_state()
        moved = state.archive_history()
        state.save()
    return {"ok": True, "archived": moved}


@app.post("/api/history/archive-item")
async def api_archive_history_item(request: Request) -> dict[str, Any]:
    protected_json(request)
    form = await request.form()
    article = str(form.get("article", ""))
    brand = str(form.get("brand", ""))
    updated_at = str(form.get("updated_at", ""))
    with STATE_LOCK:
        state = get_state()
        archived = state.archive_history_item(article, updated_at, brand)
        state.save()
    if not archived:
        raise HTTPException(status_code=404, detail="Подію не знайдено в останніх подіях.")
    return {"ok": True, "archived": 1}


def configure_runtime(config_file: Path) -> None:
    global CONFIG_FILE, RAW_CONFIG, XML_CONFIG, HOROSHOP_SETTINGS, SERVER_SETTINGS, STATE
    CONFIG_FILE = config_file
    configure_console_encoding()
    RAW_CONFIG = read_raw_config(config_file)
    XML_CONFIG = load_config(config_file)
    HOROSHOP_SETTINGS = load_horoshop_settings(RAW_CONFIG)
    SERVER_SETTINGS = load_server_settings(RAW_CONFIG, XML_CONFIG)
    initialize_reliable_clock()
    configure_logging(XML_CONFIG)
    with STATE_LOCK:
        STATE = SyncState(SERVER_SETTINGS["state_file"])
    log(
        "Horoshop Sync ініціалізовано: "
        f"порт {SERVER_SETTINGS['port']}, галерея {HOROSHOP_SETTINGS.image_field}."
    )


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Локальна веб-панель синхронізації зображень із Хорошопом."
    )
    parser.add_argument(
        "--config",
        type=Path,
        default=DEFAULT_CONFIG_FILE,
        help="Шлях до config.json.",
    )
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    try:
        configure_runtime(args.config.resolve())
    except Exception as error:
        print(f"Помилка конфігурації: {error}", flush=True)
        return 1

    uvicorn.run(
        app,
        host=SERVER_SETTINGS["host"],
        port=SERVER_SETTINGS["port"],
        log_level="info",
        access_log=True,
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
